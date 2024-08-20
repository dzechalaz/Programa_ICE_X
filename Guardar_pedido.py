import openpyxl
import qrcode
from PIL import Image
import time
import pywhatkit
import pyautogui as pg

def obtener_numero_pedido():
    archivo_excel = "Clientes.xlsx"
    nombre_hoja = "Ventas"
    
    try:
        libro_excel = openpyxl.load_workbook(archivo_excel)
        hoja = libro_excel[nombre_hoja]
        
        numero_pedido = 0
        for row in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, int):
                    numero_pedido = max(numero_pedido, cell.value)
        libro_excel.close()
        return numero_pedido + 1  # Incrementa el número para el nuevo pedido
    except FileNotFoundError:
        return 1  # Si el archivo no existe, retorna 1 (primer pedido)
    

def obtener_primera_fila_vacia(hoja):
    for fila in range(2, hoja.max_row + 2):
        if hoja.cell(row=fila, column=1).value is None:  # Solo verifica la primera columna de cada fila
            return fila
    return hoja.max_row + 1

def guardar_pedido_excel(pedidos, costo_total, pago_total, devolucion_total):
    archivo_excel = "Clientes.xlsx"
    nombre_hoja = "Ventas"

    try:
        libro_excel = openpyxl.load_workbook(archivo_excel)
    except FileNotFoundError:
        libro_excel = openpyxl.Workbook()
        libro_excel.create_sheet(title=nombre_hoja, index=0)

    hoja = libro_excel[nombre_hoja]

    encabezados = ["Pedido", "Producto", "Costo", "Total", "Pago", "Devolución"]
    if hoja.max_row == 1 and not any(cell.value for cell in hoja[1]):
        hoja.append(encabezados)
    numero_pedido = obtener_numero_pedido()

    nueva_fila = [
        numero_pedido,  
        pedidos,
        1,
        costo_total,
        pago_total,
        devolucion_total
    ]
    
    primera_fila_vacia = obtener_primera_fila_vacia(hoja)
    for col_num, value in enumerate(nueva_fila, 1):
        hoja.cell(row=primera_fila_vacia, column=col_num, value=value)
    
    libro_excel.save(archivo_excel)
    libro_excel.close()
    print("Datos guardados")

# 
# def guardar_pedido_excel(pedidos, costo_total, pago_total, devolucion_total):
#     archivo_excel = "Clientes 1.xlsx"
#     nombre_hoja = "Ventas"

#     try:
#         libro_excel = openpyxl.load_workbook(archivo_excel)
#     except FileNotFoundError:
#         libro_excel = openpyxl.Workbook()
#         libro_excel.create_sheet(title=nombre_hoja, index=0)

#     hoja = libro_excel[nombre_hoja]

#     encabezados = ["Pedido", "Producto", "Costo", "Total", "Pago", "Devolución"]
#     if hoja.max_row == 1 and not any(cell.value for cell in hoja[1]):
#         hoja.append(encabezados)

#     primera_fila_vacia = obtener_primera_fila_vacia(hoja)
#     numero_pedido = obtener_numero_pedido()
    
#     nueva_fila = [
#         numero_pedido,  
#         pedidos,
#         1,
#         costo_total,
#         pago_total,
#         devolucion_total
#     ]
#     hoja.append(nueva_fila)

#     libro_excel.save(archivo_excel)

# 

def guardar_devolucion_excel(nombre, telefono, costo_total, pago_total, devolucion_total, numero_pedido):
    archivo_excel = "Clientes 1.xlsx"
    nombre_hoja = "Devolución"

    try:
        libro_excel = openpyxl.load_workbook(archivo_excel)
    except FileNotFoundError:
        libro_excel = openpyxl.Workbook()
        libro_excel.create_sheet(title=nombre_hoja, index=0)

    hoja = libro_excel[nombre_hoja]

    encabezados = ["Pedido", "Nombre", "Celular", "Total", "Pago", "Devolución", "Folio"]
    if hoja.max_row == 1 and not any(cell.value for cell in hoja[1]):
        hoja.append(encabezados)

    primera_fila_vacia = obtener_primera_fila_vacia(hoja)

    id_unico = format(id(telefono), 'x')
    
    nueva_fila_devolucion = [
        numero_pedido,  
        nombre,
        f"+52{telefono}",
        costo_total,
        pago_total,
        devolucion_total,
        id_unico
    ]
    for col_num, value in enumerate(nueva_fila_devolucion, 1):
        hoja.cell(row=primera_fila_vacia, column=col_num, value=value)

    libro_excel.save(archivo_excel)
    libro_excel.close()

def generar_qr_con_logo(nombre, id_unico, numero_pedido, costo_total, pago_total, devolucion_total, telefono):
    # Configuración del logo
    Logo_link = 'LogoX.jpg'
    logo = Image.open(Logo_link)
 
    # Ajustar tamaño del logo
    basewidth = 90
    wpercent = (basewidth / float(logo.size[0]))
    hsize = int((float(logo.size[1]) * float(wpercent)))
    logo = logo.resize((basewidth, hsize))
    
    # Crear el QR Code
    QRcode = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_H
    )
    
    # Añadir datos (id_unico) al QR Code
    QRcode.add_data(id_unico)
    QRcode.make()
    
    # Configurar color del QR Code
    QRcolor = 'Black'
    QRimg = QRcode.make_image(
        fill_color=QRcolor, back_color="white").convert('RGB')
    
    # Posicionar logo en el QR Code
    pos = ((QRimg.size[0] - logo.size[0]) // 2,
           (QRimg.size[1] - logo.size[1]) // 2)
    QRimg.paste(logo, pos)
    
    # Guardar imagen del QR
    nombre_archivo = f'{numero_pedido}_QR.png'
    QRimg.save(nombre_archivo)

    if pago_total >= costo_total:
        devolucion_total = pago_total - costo_total

    # Guardar pedido en Excel
    guardar_pedido_excel(nombre, costo_total, pago_total, devolucion_total, numero_pedido)

    if devolucion_total > 0:
        # Guardar devolución en Excel
        guardar_devolucion_excel(nombre, telefono, costo_total, pago_total, devolucion_total, numero_pedido)
        id_unico = format(id(telefono), 'x')
        generar_qr_con_logo(nombre, id_unico, numero_pedido, costo_total, pago_total, devolucion_total, telefono)
        telefono_whats = "52" + telefono
        # Generar mensaje de WhatsApp personalizado
        mensaje_whats = f"¡Hola, {nombre}! Gracias por comprar con nosotros 🙌. Tienes una devolución de ${devolucion_total}."
        # Enviar whatsapp con mensaje y QR
        pywhatkit.sendwhats_image(telefono_whats, f"{numero_pedido}_QR.png", caption=mensaje_whats)
        time.sleep(3)           # Esperar 3 segundos a que se envíe el mensaje
        pg.hotkey('ctrl', 'w')  # Cerrar la pestaña
